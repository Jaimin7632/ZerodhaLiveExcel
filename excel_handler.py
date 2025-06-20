# live_tick_excel_app/excel_handler.py
from openpyxl import load_workbook, Workbook
from config import EXCEL_INPUT_FILE_PATH, EXCEL_INPUT_SHEET_NAME, EXCEL_INPUT_HEADER, EXCEL_INPUT_DATA_START_ROW
import os


def create_empty_excel_with_headers():
    """Creates an empty Excel file with predefined headers for instrument input."""
    wb = Workbook()
    ws = wb.active
    ws.title = EXCEL_INPUT_SHEET_NAME
    ws.append(EXCEL_INPUT_HEADER)  # Write headers to the first row
    try:
        wb.save(EXCEL_INPUT_FILE_PATH)
        print(f"[Excel Handler] Created empty Excel file: '{EXCEL_INPUT_FILE_PATH}' with headers.")
        print(
            f"[Excel Handler] Please add your 'Exchange' and 'Instrument_Name' pairs starting from row {EXCEL_INPUT_DATA_START_ROW}.")
        return True
    except Exception as e:
        print(f"[Excel Handler] Error creating Excel file '{EXCEL_INPUT_FILE_PATH}': {e}")
        return False


def read_instrument_list_from_excel():
    """Reads the list of instruments (Exchange, Instrument_Name) from the Excel file."""
    instruments_to_find = []  # List of (exchange, instrument_name) tuples

    if not os.path.exists(EXCEL_INPUT_FILE_PATH):
        print(f"[Excel Handler] Excel input file '{EXCEL_INPUT_FILE_PATH}' not found. Attempting to create.")
        if create_empty_excel_with_headers():
            print(f"[Excel Handler] Please populate '{EXCEL_INPUT_FILE_PATH}' and restart the application.")
        return []  # Return empty list, and prompt user to fill the file

    try:
        workbook = load_workbook(EXCEL_INPUT_FILE_PATH,
                                 data_only=True)  # data_only=True to read cell values, not formulas
        sheet = workbook[EXCEL_INPUT_SHEET_NAME]

        # Basic header validation (optional but good practice)
        header_row = [cell.value for cell in sheet[1]]
        if header_row[:len(EXCEL_INPUT_HEADER)] != EXCEL_INPUT_HEADER:
            print(
                f"[Excel Handler] Warning: Excel headers mismatch. Expected {EXCEL_INPUT_HEADER}, got {header_row[:len(EXCEL_INPUT_HEADER)]}. Proceeding anyway.")

        for row_index in range(EXCEL_INPUT_DATA_START_ROW, sheet.max_row + 1):
            # exchange_cell_val = sheet.cell(row=row_index, column=1).value
            instrument_name_cell_val = sheet.cell(row=row_index, column=1).value

            if instrument_name_cell_val:
                instruments_to_find.append(str(instrument_name_cell_val).strip().upper())

        if not instruments_to_find:
            print(
                f"[Excel Handler] No instruments found in the '{EXCEL_INPUT_SHEET_NAME}' sheet of '{EXCEL_INPUT_FILE_PATH}'.")
            return []

        print(f"[Excel Handler] Read {len(instruments_to_find)} entries from '{EXCEL_INPUT_FILE_PATH}'.")
        return instruments_to_find

    except KeyError:
        print(f"[Excel Handler] Sheet '{EXCEL_INPUT_SHEET_NAME}' not found in '{EXCEL_INPUT_FILE_PATH}'.")
        print(f"[Excel Handler] Ensure your sheet is named '{EXCEL_INPUT_SHEET_NAME}' or update config.py.")
        return []
    except Exception as e:
        print(f"[Excel Handler] Error reading Excel file: {e}")
        return []